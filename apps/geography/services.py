from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO, TextIO

from openpyxl import load_workbook

NORMALIZED_COLUMNS = (
    "department_code",
    "department_name",
    "province_code",
    "province_name",
    "district_code",
    "district_name",
)
OFFICIAL_RESOURCE_URL = (
    "https://cdn.www.gob.pe/uploads/document/file/8261096/"
    "6894980-peru-poblacion-total-proyectada-al-30-de-junio-de-cada-ano-"
    "segun-departamento-provincia-y-distrito-2018-2026.xlsx?v=1768402069"
)
UBIGEO_PATTERN = re.compile(r"[0-9]{6}")
FOOTNOTE_SUFFIX = re.compile(r"(?:\s+\d+/)+\s*$")


class GeographySourceError(ValueError):
    pass


@dataclass(frozen=True, order=True)
class DepartmentRecord:
    code: str
    name: str


@dataclass(frozen=True, order=True)
class ProvinceRecord:
    code: str
    name: str
    department_code: str


@dataclass(frozen=True, order=True)
class DistrictRecord:
    code: str
    name: str
    province_code: str


@dataclass(frozen=True)
class GeographySnapshot:
    departments: tuple[DepartmentRecord, ...]
    provinces: tuple[ProvinceRecord, ...]
    districts: tuple[DistrictRecord, ...]


def normalize_official_name(value: object) -> str:
    text = " ".join(str(value or "").strip().split())
    text = FOOTNOTE_SUFFIX.sub("", text)
    words = text.title().split()
    return " ".join(
        word.lower() if index and word.lower() in {"de", "del", "y"} else word
        for index, word in enumerate(words)
    )


def parse_geography_source(source: str | Path | BinaryIO) -> GeographySnapshot:
    if hasattr(source, "read"):
        return _parse_xlsx(source)
    path = Path(source)
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8", newline="") as source_file:
            return _parse_csv(source_file)
    with path.open("rb") as source_file:
        return _parse_xlsx(source_file)


def parse_geography_bytes(raw: bytes) -> GeographySnapshot:
    return _parse_xlsx(io.BytesIO(raw))


def _parse_xlsx(source: BinaryIO) -> GeographySnapshot:
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as error:
        raise GeographySourceError(
            f"The XLSX source could not be read: {error}"
        ) from error

    try:
        sheet = workbook["POB. PROYECTADA 2018-2026"]
    except KeyError as error:
        raise GeographySourceError(
            "The expected INEI worksheet 'POB. PROYECTADA 2018-2026' is absent."
        ) from error

    departments: list[DepartmentRecord] = []
    provinces: list[ProvinceRecord] = []
    districts: list[DistrictRecord] = []
    for row in sheet.iter_rows(values_only=True):
        raw_code = row[0] if row else None
        raw_name = row[1] if len(row) > 1 else None
        if isinstance(raw_code, int | float):
            code = str(int(raw_code)).zfill(6)
        else:
            code = str(raw_code or "").strip()
        if not UBIGEO_PATTERN.fullmatch(code) or code == "000000":
            continue
        name = normalize_official_name(raw_name)
        if code.endswith("0000"):
            departments.append(DepartmentRecord(code[:2], name))
        elif code.endswith("00"):
            provinces.append(ProvinceRecord(code[:4], name, code[:2]))
        else:
            districts.append(DistrictRecord(code, name, code[:4]))
    return validate_snapshot(departments, provinces, districts)


def _parse_csv(source: TextIO) -> GeographySnapshot:
    reader = csv.DictReader(source)
    if tuple(reader.fieldnames or ()) != NORMALIZED_COLUMNS:
        raise GeographySourceError(
            "The normalized CSV columns must be exactly: "
            + ", ".join(NORMALIZED_COLUMNS)
        )

    department_rows: dict[str, DepartmentRecord] = {}
    province_rows: dict[str, ProvinceRecord] = {}
    districts: list[DistrictRecord] = []
    for line_number, row in enumerate(reader, start=2):
        values = {key: str(row.get(key, "")).strip() for key in NORMALIZED_COLUMNS}
        if not all(values.values()):
            raise GeographySourceError(f"CSV row {line_number} contains blank values.")
        department = DepartmentRecord(
            values["department_code"], values["department_name"]
        )
        province = ProvinceRecord(
            values["province_code"],
            values["province_name"],
            values["department_code"],
        )
        previous_department = department_rows.setdefault(department.code, department)
        if previous_department != department:
            raise GeographySourceError(
                f"Department {department.code} has conflicting rows."
            )
        previous_province = province_rows.setdefault(province.code, province)
        if previous_province != province:
            raise GeographySourceError(
                f"Province {province.code} has conflicting rows."
            )
        districts.append(
            DistrictRecord(
                values["district_code"],
                values["district_name"],
                values["province_code"],
            )
        )
    return validate_snapshot(
        department_rows.values(), province_rows.values(), districts
    )


def validate_snapshot(departments, provinces, districts) -> GeographySnapshot:
    department_rows = tuple(sorted(departments))
    province_rows = tuple(sorted(provinces))
    district_rows = tuple(sorted(districts))
    _validate_records(department_rows, 2, "department")
    _validate_records(province_rows, 4, "province")
    _validate_records(district_rows, 6, "district")

    department_codes = {row.code for row in department_rows}
    province_codes = {row.code for row in province_rows}
    orphan_provinces = [
        row.code
        for row in province_rows
        if row.department_code not in department_codes
        or not row.code.startswith(row.department_code)
    ]
    orphan_districts = [
        row.code
        for row in district_rows
        if row.province_code not in province_codes
        or not row.code.startswith(row.province_code)
    ]
    if orphan_provinces:
        raise GeographySourceError("Orphan provinces: " + ", ".join(orphan_provinces))
    if orphan_districts:
        raise GeographySourceError("Orphan districts: " + ", ".join(orphan_districts))
    return GeographySnapshot(department_rows, province_rows, district_rows)


def _validate_records(records, code_length: int, label: str) -> None:
    codes: set[str] = set()
    for row in records:
        if len(row.code) != code_length or not row.code.isdigit():
            raise GeographySourceError(f"Malformed {label} code: {row.code!r}")
        if not row.name:
            raise GeographySourceError(f"{label.title()} {row.code} has no name.")
        if row.code in codes:
            raise GeographySourceError(f"Duplicate {label} code: {row.code}")
        codes.add(row.code)
